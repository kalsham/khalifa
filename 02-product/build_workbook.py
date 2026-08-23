#!/usr/bin/env python3
"""Build the UGC Creator Business OS workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

FONT = "Arial"

# Palette: off-black / cream / coral accent / sage (positive) / dusty rose (negative)
INK = "1C1C1E"
CREAM = "FAF6EE"
CARD = "F1EAD9"
ACCENT = "E8674A"
SAGE = "8FA98B"
ROSE = "D98B87"
GOLD = "D4A853"
LAVENDER = "C9B8D8"
SKYBLUE = "AFC8D6"
GRAY_TXT = "6B6B66"
WHITE = "FFFFFF"

def hx(c):
    return c if c.startswith("00") or len(c) == 8 else c

def fill(color):
    return PatternFill("solid", fgColor=color)

def thin_border(color="D8D0BE"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

wb = openpyxl.Workbook()

# ---------------------------------------------------------------- Dashboard
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGH", [3, 20, 20, 20, 20, 20, 20, 3]):
    ws.column_dimensions[col].width = w

ws.merge_cells("B1:G1")
ws["B1"] = "UGC CREATOR BUSINESS OS"
ws["B1"].font = Font(name=FONT, size=22, bold=True, color=WHITE)
ws["B1"].fill = fill(INK)
ws["B1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 42

ws.merge_cells("B2:G2")
ws["B2"] = "Your brand-deal command center — pipeline, pricing, invoices, and content in one place."
ws["B2"].font = Font(name=FONT, size=11, italic=True, color=GRAY_TXT)
ws["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 22

def stat_card(cell_range_label, cell_range_value, label, formula, fmt, color):
    ws.merge_cells(cell_range_label)
    lc = cell_range_label.split(":")[0]
    ws[lc] = label
    ws[lc].font = Font(name=FONT, size=10, bold=True, color=WHITE)
    ws[lc].fill = fill(color)
    ws[lc].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(cell_range_value)
    vc = cell_range_value.split(":")[0]
    ws[vc] = formula
    ws[vc].number_format = fmt
    ws[vc].font = Font(name=FONT, size=18, bold=True, color=INK)
    ws[vc].fill = fill(CARD)
    ws[vc].alignment = Alignment(horizontal="center", vertical="center")

for rng in ["B4:C4", "D4:E4", "F4:G4", "B5:C6", "D5:E6", "F5:G6",
            "B8:C8", "D8:E8", "F8:G8", "B9:C10", "D9:E10", "F9:G10"]:
    pass  # placeholder rows sized below

ws.row_dimensions[4].height = 20
ws.row_dimensions[5].height = 28
ws.row_dimensions[6].height = 10
ws.row_dimensions[8].height = 20
ws.row_dimensions[9].height = 28
ws.row_dimensions[10].height = 10

CUR = '$#,##0.00'
PCT = '0.0%'
NUM = '#,##0'

stat_card("B4:C4", "B5:C6", "ACTIVE PIPELINE VALUE",
          "=SUMPRODUCT((('Brand Deal Pipeline'!$E$4:$E$103=\"Lead\")+('Brand Deal Pipeline'!$E$4:$E$103=\"Pitched\")+('Brand Deal Pipeline'!$E$4:$E$103=\"Negotiating\")+('Brand Deal Pipeline'!$E$4:$E$103=\"Contract Sent\"))*('Brand Deal Pipeline'!$I$4:$I$103))",
          CUR, INK)
stat_card("D4:E4", "D5:E6", "REVENUE — THIS MONTH",
          "=SUMIFS(Invoices!$E$4:$E$103,Invoices!$F$4:$F$103,\"Paid\",Invoices!$I$4:$I$103,\">=\"&$B$31,Invoices!$I$4:$I$103,\"<=\"&$C$31)",
          CUR, ACCENT)
stat_card("F4:G4", "F5:G6", "REVENUE — THIS QUARTER",
          "=SUMIFS(Invoices!$E$4:$E$103,Invoices!$F$4:$F$103,\"Paid\",Invoices!$I$4:$I$103,\">=\"&$B$32,Invoices!$I$4:$I$103,\"<=\"&$C$32)",
          CUR, ACCENT)
stat_card("B8:C8", "B9:C10", "REVENUE — THIS YEAR",
          "=SUMIFS(Invoices!$E$4:$E$103,Invoices!$F$4:$F$103,\"Paid\",Invoices!$I$4:$I$103,\">=\"&$B$33,Invoices!$I$4:$I$103,\"<=\"&$C$33)",
          CUR, SAGE)
stat_card("D8:E8", "D9:E10", "ACTIVE DEALS",
          "=SUMPRODUCT((('Brand Deal Pipeline'!$E$4:$E$103=\"Lead\")+('Brand Deal Pipeline'!$E$4:$E$103=\"Pitched\")+('Brand Deal Pipeline'!$E$4:$E$103=\"Negotiating\")+('Brand Deal Pipeline'!$E$4:$E$103=\"Contract Sent\"))*1)",
          NUM, GOLD)
stat_card("F8:G8", "F9:G10", "WIN RATE",
          "=IFERROR(COUNTIF('Brand Deal Pipeline'!$E$4:$E$103,\"Closed-Won\")/(COUNTIF('Brand Deal Pipeline'!$E$4:$E$103,\"Closed-Won\")+COUNTIF('Brand Deal Pipeline'!$E$4:$E$103,\"Closed-Lost\")),0)",
          PCT, GOLD)

ws.row_dimensions[12].height = 20
ws.row_dimensions[13].height = 28
stat_card("B12:C12", "B13:C14", "AVG DEAL SIZE (CLOSED-WON)",
          "=IFERROR(AVERAGEIF('Brand Deal Pipeline'!$E$4:$E$103,\"Closed-Won\",'Brand Deal Pipeline'!$I$4:$I$103),0)",
          CUR, LAVENDER)
stat_card("D12:E12", "D13:E14", "TOTAL UNPAID INVOICES",
          "=SUMIF(Invoices!$F$4:$F$103,\"Unpaid\",Invoices!$E$4:$E$103)",
          CUR, ROSE)
stat_card("F12:G12", "F13:G14", "TOTAL PAID INVOICES",
          "=SUMIF(Invoices!$F$4:$F$103,\"Paid\",Invoices!$E$4:$E$103)",
          CUR, SAGE)

ws.merge_cells("B16:G16")
ws["B16"] = "How to use this workbook"
ws["B16"].font = Font(name=FONT, size=12, bold=True, color=INK)
tips = [
    "1. Brand Deal Pipeline — log every brand conversation the moment it starts, even a cold DM.",
    "2. Rate Calculator — use it BEFORE you quote a brand. Never quote off the top of your head.",
    "3. Invoices — every deal that closes gets an invoice row, so revenue here stays real.",
    "4. Content Calendar — plan filming/editing/posting per deal so nothing misses a deadline.",
    "5. Usage Rights Cheat Sheet — read this once. It's the #1 place new UGC creators lose money.",
]
for i, t in enumerate(tips):
    r = 17 + i
    ws.merge_cells(f"B{r}:G{r}")
    ws[f"B{r}"] = t
    ws[f"B{r}"].font = Font(name=FONT, size=10, color=GRAY_TXT)
    ws.row_dimensions[r].height = 16

# internal date-boundary helpers
ws["B29"] = "Internal helper cells — power the date formulas above. Don't delete."
ws["B29"].font = Font(name=FONT, size=8, italic=True, color=GRAY_TXT)
ws.merge_cells("B29:G29")
ws["A30"] = "Range"
ws["B30"] = "Start"
ws["C30"] = "End"
for c in ["A30", "B30", "C30"]:
    ws[c].font = Font(name=FONT, size=8, bold=True, color=GRAY_TXT)
ws["A31"] = "Month"; ws["B31"] = "=DATE(YEAR(TODAY()),MONTH(TODAY()),1)"; ws["C31"] = "=EOMONTH(B31,0)"
ws["A32"] = "Quarter"; ws["B32"] = "=DATE(YEAR(TODAY()),(ROUNDUP(MONTH(TODAY())/3,0)*3)-2,1)"; ws["C32"] = "=EOMONTH(B32,2)"
ws["A33"] = "Year"; ws["B33"] = "=DATE(YEAR(TODAY()),1,1)"; ws["C33"] = "=DATE(YEAR(TODAY()),12,31)"
for r in range(31, 34):
    for c in ["A", "B", "C"]:
        ws[f"{c}{r}"].font = Font(name=FONT, size=8, color=GRAY_TXT)
        ws[f"{c}{r}"].number_format = "yyyy-mm-dd" if c != "A" else "General"

for row in ws.iter_rows(min_row=1, max_row=34, min_col=1, max_col=8):
    for cell in row:
        if cell.fill.fgColor.rgb in (None, "00000000") :
            cell.fill = fill(CREAM)

# ------------------------------------------------------------- shared helpers
DELIVERABLE_TYPES = "Video,Photo,Video+Photo Bundle,UGC Ad Script,Whitelisted Content"
STAGES = "Lead,Pitched,Negotiating,Contract Sent,Closed-Won,Closed-Lost"
USAGE_RIGHTS = "Organic Only,Whitelisted 30 Days,Whitelisted 90 Days,Paid Ads/Whitelisted 6mo,Full Buyout"
CAL_STATUS = "Not Started,Filming,Editing,Submitted,Posted"
INV_STATUS = "Unpaid,Paid,Overdue"

STAGE_COLORS = {
    "Lead": SKYBLUE, "Pitched": GOLD, "Negotiating": "F2D93A",
    "Contract Sent": LAVENDER, "Closed-Won": SAGE, "Closed-Lost": ROSE,
}

def title_row(sheet, text, subtitle, ncols):
    last = get_column_letter(ncols)
    sheet.merge_cells(f"A1:{last}1")
    sheet["A1"] = text
    sheet["A1"].font = Font(name=FONT, size=16, bold=True, color=WHITE)
    sheet["A1"].fill = fill(INK)
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells(f"A2:{last}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name=FONT, size=9, italic=True, color=GRAY_TXT)
    sheet["A2"].fill = fill(CARD)
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sheet.row_dimensions[2].height = 18

def header_row(sheet, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = sheet.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
        c.fill = fill(INK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.row_dimensions[row].height = 30

def example_row(sheet, row, values, ncols):
    for i, v in enumerate(values, start=1):
        c = sheet.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=10, italic=True, color=GRAY_TXT)
        c.fill = fill(CARD)
        c.border = thin_border()
    last = get_column_letter(ncols)

def add_dropdown(sheet, col_letter, first, last, options_csv):
    dv = DataValidation(type="list", formula1=f'"{options_csv}"', allow_blank=True, showDropDown=False)
    sheet.add_data_validation(dv)
    dv.add(f"{col_letter}{first}:{col_letter}{last}")

def style_blank_rows(sheet, first, last, ncols, alt=False):
    for r in range(first, last + 1):
        for c in range(1, ncols + 1):
            cell = sheet.cell(row=r, column=c)
            cell.border = thin_border()
            cell.font = Font(name=FONT, size=10, color=INK)
            if alt and r % 2 == 0:
                cell.fill = fill(CREAM)
            else:
                cell.fill = fill(WHITE)

FIRST_DATA = 4
LAST_DATA = 103

# ------------------------------------------------------- Brand Deal Pipeline
p = wb.create_sheet("Brand Deal Pipeline")
p.sheet_view.showGridLines = False
title_row(p, "BRAND DEAL PIPELINE", "Log every deal the moment it starts. Row 3 is an example — overwrite it.", 11)
headers = ["Brand Name", "Contact Name", "Contact Email", "Deliverable Type", "Stage",
           "Rate Quoted", "Usage Rights", "Usage Fee Add-on", "Total Deal Value", "Deadline", "Notes"]
widths = [18, 16, 22, 18, 14, 12, 20, 14, 14, 12, 26]
header_row(p, 3, headers, widths)
example_row(p, 4, ["Glow Skincare Co.", "Maria Lopez", "maria@glowskincare.com", "Video",
                    "Negotiating", 250, "Whitelisted 30 Days", 62.5, "=F4+H4",
                    "2026-09-15", "Wants 2 revisions included"], 11)
style_blank_rows(p, 5, LAST_DATA, 11, alt=True)
for r in range(5, LAST_DATA + 1):
    p.cell(row=r, column=9, value=f"=F{r}+H{r}")
    p.cell(row=r, column=9).number_format = CUR
p.cell(row=4, column=9).number_format = CUR
p.cell(row=4, column=6).number_format = CUR
p.cell(row=4, column=8).number_format = CUR
for r in range(4, LAST_DATA + 1):
    p.cell(row=r, column=6).number_format = CUR
    p.cell(row=r, column=8).number_format = CUR
    p.cell(row=r, column=10).number_format = "yyyy-mm-dd"
add_dropdown(p, "D", 4, LAST_DATA, DELIVERABLE_TYPES)
add_dropdown(p, "E", 4, LAST_DATA, STAGES)
add_dropdown(p, "G", 4, LAST_DATA, USAGE_RIGHTS)
for stage, color in STAGE_COLORS.items():
    p.conditional_formatting.add(
        f"E4:E{LAST_DATA}",
        CellIsRule(operator="equal", formula=[f'"{stage}"'], fill=fill(color))
    )
p.freeze_panes = "A4"

# ----------------------------------------------------------- Rate Calculator
rc = wb.create_sheet("Rate Calculator")
rc.sheet_view.showGridLines = False
title_row(rc, "RATE CALCULATOR", "Use this BEFORE you quote a brand. Never price off the top of your head.", 4)
for col, w in zip("ABCD", [30, 18, 18, 40]):
    rc.column_dimensions[col].width = w

rc["A4"] = "STEP 1 — Your base hourly-equivalent rate"
rc["A4"].font = Font(name=FONT, size=12, bold=True, color=INK)
rc.merge_cells("A4:D4")
rc["A5"] = "Base rate ($/hour)"
rc["A5"].font = Font(name=FONT, size=10)
rc["B5"] = 50
rc["B5"].font = Font(name=FONT, size=12, bold=True, color="0000FF")
rc["B5"].fill = fill("FFFDE7")
rc["B5"].number_format = CUR
rc["B5"].border = thin_border()
rc["D5"] = "← edit this. Blue cells are yours to change."
rc["D5"].font = Font(name=FONT, size=9, italic=True, color=GRAY_TXT)

rc["A7"] = "STEP 2 — Estimated hours per deliverable (edit to match your speed)"
rc["A7"].font = Font(name=FONT, size=12, bold=True, color=INK)
rc.merge_cells("A7:D7")
header_row(rc, 8, ["Deliverable Type", "Est. Hours", "", ""], [30, 18, 18, 40])
hours_defaults = [("Video", 3), ("Photo", 1), ("Video+Photo Bundle", 4),
                   ("UGC Ad Script", 2), ("Whitelisted Content", 3)]
for i, (dtype, hrs) in enumerate(hours_defaults):
    r = 9 + i
    rc.cell(row=r, column=1, value=dtype).font = Font(name=FONT, size=10)
    c = rc.cell(row=r, column=2, value=hrs)
    c.font = Font(name=FONT, size=10, color="0000FF")
    c.fill = fill("FFFDE7")
    c.number_format = "0.0"
    for col in (1, 2):
        rc.cell(row=r, column=col).border = thin_border()

rc["A15"] = "STEP 3 — Usage rights multiplier (edit if your market is different)"
rc["A15"].font = Font(name=FONT, size=12, bold=True, color=INK)
rc.merge_cells("A15:D15")
header_row(rc, 16, ["Usage Rights", "Multiplier", "", ""], [30, 18, 18, 40])
mult_defaults = [("Organic Only", 1.0), ("Whitelisted 30 Days", 1.25), ("Whitelisted 90 Days", 1.5),
                  ("Paid Ads/Whitelisted 6mo", 2.0), ("Full Buyout", 3.0)]
for i, (tier, m) in enumerate(mult_defaults):
    r = 17 + i
    rc.cell(row=r, column=1, value=tier).font = Font(name=FONT, size=10)
    c = rc.cell(row=r, column=2, value=m)
    c.font = Font(name=FONT, size=10, color="0000FF")
    c.fill = fill("FFFDE7")
    c.number_format = "0.00x"
    for col in (1, 2):
        rc.cell(row=r, column=col).border = thin_border()

rc["A23"] = "CALCULATE YOUR RATE"
rc["A23"].font = Font(name=FONT, size=13, bold=True, color=WHITE)
rc.merge_cells("A23:D23")
rc["A23"].fill = fill(ACCENT)
rc.row_dimensions[23].height = 24

rc["A24"] = "Deliverable Type"
rc["B24"] = "Video"
rc["B24"].fill = fill("FFFDE7")
rc["B24"].border = thin_border()
add_dropdown(rc, "B", 24, 24, DELIVERABLE_TYPES)

rc["A25"] = "Usage Rights"
rc["B25"] = "Organic Only"
rc["B25"].fill = fill("FFFDE7")
rc["B25"].border = thin_border()
add_dropdown(rc, "B", 25, 25, USAGE_RIGHTS)

for r in (24, 25):
    rc.cell(row=r, column=1).font = Font(name=FONT, size=10, bold=True)

rc["A27"] = "Matched hours (from Step 2)"
rc["B27"] = '=INDEX($B$9:$B$13,MATCH($B$24,$A$9:$A$13,0))'
rc["A28"] = "Base deliverable price (rate x hours)"
rc["B28"] = "=$B$5*$B$27"
rc["B28"].number_format = CUR
rc["A29"] = "Matched multiplier (from Step 3)"
rc["B29"] = '=INDEX($B$17:$B$21,MATCH($B$25,$A$17:$A$21,0))'
rc["A30"] = "SUGGESTED RATE"
rc["A30"].font = Font(name=FONT, size=13, bold=True, color=INK)
rc["B30"] = "=ROUND($B$28*$B$29,0)"
rc["B30"].number_format = CUR
rc["B30"].font = Font(name=FONT, size=18, bold=True, color=INK)
rc["B30"].fill = fill(CARD)
rc.merge_cells("B30:C30")
for r in (27, 28, 29):
    rc.cell(row=r, column=1).font = Font(name=FONT, size=10, color=GRAY_TXT)
    rc.cell(row=r, column=2).font = Font(name=FONT, size=10, color=INK)
rc["D28"] = "= Base Rate x Hours"
rc["D30"] = "= Base Deliverable Price x Multiplier. This is your floor, not your ceiling — negotiate up."
rc["D30"].font = Font(name=FONT, size=9, italic=True, color=GRAY_TXT)
rc["D28"].font = Font(name=FONT, size=9, italic=True, color=GRAY_TXT)

for row in rc.iter_rows(min_row=1, max_row=32, min_col=1, max_col=4):
    for cell in row:
        if cell.fill.fgColor.rgb in (None, "00000000"):
            cell.fill = fill(CREAM)

# ------------------------------------------------------------------ Invoices
inv = wb.create_sheet("Invoices")
inv.sheet_view.showGridLines = False
title_row(inv, "INVOICES", "Every closed deal gets a row here — this is what powers the Dashboard revenue numbers.", 9)
headers = ["Invoice #", "Brand Name", "Date Issued", "Due Date", "Amount",
           "Status", "Auto-Flag", "Payment Method", "Date Paid"]
widths = [12, 20, 14, 14, 14, 12, 14, 16, 14]
header_row(inv, 3, headers, widths)
example_row(inv, 4, ["INV-0001", "Glow Skincare Co.", "2026-08-01", "2026-08-15", 312.5,
                      "Paid", "", "PayPal", "2026-08-10"], 9)
style_blank_rows(inv, 5, LAST_DATA, 9, alt=True)
for r in range(4, LAST_DATA + 1):
    inv.cell(row=r, column=3).number_format = "yyyy-mm-dd"
    inv.cell(row=r, column=4).number_format = "yyyy-mm-dd"
    inv.cell(row=r, column=5).number_format = CUR
    inv.cell(row=r, column=9).number_format = "yyyy-mm-dd"
    inv.cell(row=r, column=7, value=f'=IF(AND(F{r}<>"Paid",F{r}<>"",TODAY()>D{r}),"⚠ PAST DUE","")')
    inv.cell(row=r, column=7).font = Font(name=FONT, size=10, bold=True, color=ACCENT)
add_dropdown(inv, "F", 4, LAST_DATA, INV_STATUS)
inv.conditional_formatting.add(
    f"G4:G{LAST_DATA}",
    FormulaRule(formula=[f'$G4<>""'], fill=fill(ROSE))
)
inv.conditional_formatting.add(
    f"F4:F{LAST_DATA}",
    CellIsRule(operator="equal", formula=['"Paid"'], fill=fill(SAGE))
)
inv.conditional_formatting.add(
    f"F4:F{LAST_DATA}",
    CellIsRule(operator="equal", formula=['"Unpaid"'], fill=fill(GOLD))
)
inv.freeze_panes = "A4"

tr = LAST_DATA + 2
inv[f"B{tr}"] = "TOTAL PAID"
inv[f"B{tr}"].font = Font(name=FONT, size=11, bold=True)
inv[f"E{tr}"] = f'=SUMIF(F4:F{LAST_DATA},"Paid",E4:E{LAST_DATA})'
inv[f"E{tr}"].number_format = CUR
inv[f"E{tr}"].font = Font(name=FONT, size=11, bold=True, color=SAGE)
inv[f"B{tr+1}"] = "TOTAL UNPAID"
inv[f"B{tr+1}"].font = Font(name=FONT, size=11, bold=True)
inv[f"E{tr+1}"] = f'=SUMIF(F4:F{LAST_DATA},"Unpaid",E4:E{LAST_DATA})'
inv[f"E{tr+1}"].number_format = CUR
inv[f"E{tr+1}"].font = Font(name=FONT, size=11, bold=True, color=GOLD)

# ------------------------------------------------------------ Content Calendar
cal = wb.create_sheet("Content Calendar")
cal.sheet_view.showGridLines = False
title_row(cal, "CONTENT CALENDAR", "Plan filming, editing, and posting per deal so nothing misses a deadline.", 6)
headers = ["Date", "Brand/Deal", "Content Type", "Platform", "Status", "Caption/Notes"]
widths = [12, 20, 18, 14, 14, 34]
header_row(cal, 3, headers, widths)
example_row(cal, 4, ["2026-09-10", "Glow Skincare Co.", "Video", "TikTok", "Filming",
                      "Morning routine hook, show product 0:03"], 6)
LAST_CAL = 63
style_blank_rows(cal, 5, LAST_CAL, 6, alt=True)
for r in range(4, LAST_CAL + 1):
    cal.cell(row=r, column=1).number_format = "yyyy-mm-dd"
add_dropdown(cal, "E", 4, LAST_CAL, CAL_STATUS)
STATUS_COLORS = {"Not Started": "E6E1D3", "Filming": SKYBLUE, "Editing": GOLD,
                  "Submitted": LAVENDER, "Posted": SAGE}
for status, color in STATUS_COLORS.items():
    cal.conditional_formatting.add(
        f"E4:E{LAST_CAL}",
        CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill(color))
    )
cal.freeze_panes = "A4"

# ------------------------------------------------------ Usage Rights Cheat Sheet
ur = wb.create_sheet("Usage Rights Cheat Sheet")
ur.sheet_view.showGridLines = False
title_row(ur, "USAGE RIGHTS CHEAT SHEET", "Read this once. It's the #1 place new UGC creators leave money on the table.", 4)
for col, w in zip("ABCD", [24, 44, 16, 46]):
    ur.column_dimensions[col].width = w
header_row(ur, 3, ["Tier", "What It Actually Means", "Price Multiplier", "Use It When"], [24, 44, 16, 46])
rows = [
    ("Organic Only", "Brand only gets the raw file(s). They post nothing as an ad, and don't run it through your handle or theirs as paid media.", "1.0x", "They only want content for their own organic/unpaid social pages."),
    ("Whitelisted 30 Days", "Brand runs your content as a paid ad THROUGH YOUR HANDLE (Spark Ads / Partnership Ads) for up to 30 days.", "1.25x", "A short test campaign or one-off promo."),
    ("Whitelisted 90 Days", "Same as above, for up to 90 days — the most common ask for a standard campaign flight.", "1.5x", "A normal-length paid campaign."),
    ("Paid Ads / Whitelisted 6mo", "Extended paid-media usage, sometimes across the brand's own channels too, for up to 6 months.", "2.0x", "An always-on or longer-running paid campaign."),
    ("Full Buyout", "Brand owns unlimited, perpetual usage — ads, packaging, website, anywhere, forever.", "3.0x (floor)", "Anytime 'forever' or 'unlimited' is mentioned. Never accept 1x for this."),
]
for i, (tier, meaning, mult, when) in enumerate(rows):
    r = 4 + i
    vals = [tier, meaning, mult, when]
    for c, v in enumerate(vals, start=1):
        cell = ur.cell(row=r, column=c, value=v)
        cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin_border()
        cell.fill = fill(CREAM if r % 2 == 0 else WHITE)
    ur.row_dimensions[r].height = 44

note_row = 4 + len(rows) + 1
ur.merge_cells(f"A{note_row}:D{note_row}")
ur[f"A{note_row}"] = "Rule of thumb: if you don't know the brand's usage plan, ASK before you quote. Usage rights — not the content itself — is where creators leave the most money on the table."
ur[f"A{note_row}"].font = Font(name=FONT, size=10, bold=True, italic=True, color=ACCENT)
ur[f"A{note_row}"].alignment = Alignment(wrap_text=True, vertical="center")
ur.row_dimensions[note_row].height = 30

wb.save("/home/user/khalifa/02-product/UGC-Creator-Business-OS.xlsx")
print("workbook built")
